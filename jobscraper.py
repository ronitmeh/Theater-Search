import requests
from app_password import password
import openpyxl
import os
import sys
from urllib.request import Request, urlopen
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import bs4
from bs4 import BeautifulSoup as soup
from email.message import EmailMessage
import smtplib
import ssl
import email
from email.message import EmailMessage

def send_email(): #function which sends email to chosen recipient when new job has opened
    email_sender='theateropps@gmail.com'
    email_receiver='*****************'
    app_password = password()

    subject = 'A New Job Has Opened'
    body = 'Please see info on spreadsheet'

    em = EmailMessage()
    em["From"] = email_sender
    em['To'] = email_receiver
    em['Subject'] = subject
    em.set_content(body)

    context = ssl.create_default_context()

    attachment_path = '/Users/ronitmehta/Desktop/theatresearch/TheaterJobs.xlsm'
    with open(attachment_path, 'rb') as ap:
        em.add_attachment(ap.read(), maintype='application', subtype='vnd.ms-excel.sheet.macroEnabled.12',filename=os.path.basename(attachment_path))


    with smtplib.SMTP_SSL('smtp.gmail.com',465,context=context) as smtp:
        smtp.login(email_sender, app_password)
        smtp.sendmail(email_sender, email_receiver, em.as_string())



def theater_search(): #uses Beautiful Soup and webscraping to read through theater websites, and updates spreadsheet if new job has opened up
    new_info = False
    with open("/Users/ronitmehta/Desktop/theatresearch/Theater_Search.txt") as f:
        listoft = [line.rstrip('\n') for line in f]
    urls = ['https://keegantheatre.com/work-with-us/', 'https://theaterj.org/engage/work-with-us/#1505318914096-9ae61d0c-b186','https://www.studiotheatre.org/about/work-with-us/employment', 'https://anacostiaplayhouse.com/opportunities/', 'https://imaginationstage.org/news-events/careers/', 'https://www.nationaltheatre.org/career-opportunities/', 'https://www.shakespearetheatre.org/about/careers-and-casting/jobs/', 'https://amherst.wd5.myworkdayjobs.com/FSL_Employment_Opportunities', 'https://www.roundhousetheatre.org/About/Opportunities','https://www.woollymammoth.net/join-us/work-here/']
    theaters = ["Keegan Theatre", "Theater J", "Studio Theatre", "Anacostia Playhouse", "Imagination Stage", "National Theatre", "STC", "Folger Theatre", "Roundhouse Theatre", "Woooly Mammoth Theatre"]
    classes = ["panel-title font-weight-500 text-extra-dark-gray" ,"vc_tta-title-text", "xcms-accordion-bar", "wpgb-block-2", "header card-blog__heading", "kt-accordion-panel-inner", "text wrap", "css-1q2dra3", 'h2', "header-text"]
    os.chdir('/Users/ronitmehta/Desktop/theatresearch')
    excelFiles = os.listdir('.')
    wb = openpyxl.load_workbook(excelFiles[4], read_only=False, keep_vba=True)
    sheet = wb.active
    count = 2
    for i in range(len(urls)):
        opps = []
        u = urls[i]
        req = Request(u, headers={'User-Agent': 'Mozilla/5.0'})
        webpage = urlopen(req).read()
        website = soup(webpage, 'html.parser')
        if i != 7:
            for job in website.find_all(class_= classes[i]):
                opps.append(job.get_text().strip())
        else:
            for job in website.find_all(classes[i]):
                opps.append(job.get_text().strip())
        for opp in set(opps):
            if "Manager" in opp or "Production" in opp:
                if classes[i] == "text wrap":
                    opper = opp.split('Shake')
                    opp = opper[0]
                full_phrase = opp + ' ' + theaters[i]
                if full_phrase not in listoft:
                    with open("/Users/ronitmehta/Desktop/theatresearch/Theater_Search.txt", "a") as doc:
                        doc.write(full_phrase + '\n')


                    #if new job has opened, make cell in Excel sheet red and updates the table
                    sheet.cell(row=count, column=1).value = theaters[i]
                    sheet.cell(row=count, column=2).value = opp
                    sheet.cell(row=count, column=3).value = urls[i]
                    sheet.cell(count, 1).fill =  PatternFill(start_color='FF0000', end_color='FF0000', fill_type="solid") 
                    sheet.cell(count, 2).fill =  PatternFill(start_color='FF0000', end_color='FF0000', fill_type="solid") 
                    sheet.cell(count, 3).fill =  PatternFill(start_color='FF0000', end_color='FF0000', fill_type="solid")
                    new_info = True
                
                
                else:
                   
                    #if not a new job, add it to the sheet with no highlight
                    sheet.cell(row=count, column=1).value = theaters[i]
                    sheet.cell(row=count, column=2).value = opp
                    sheet.cell(row=count, column=3).value = urls[i]
                    sheet.cell(count, 1).fill =  PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type="solid")
                    sheet.cell(count, 2).fill =  PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type="solid")
                    sheet.cell(count, 3).fill =  PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type="solid")              
                count += 1
    wb.save(excelFiles[4])
    if new_info: #if new job, send email
        send_email()
    sys.exit()

if __name__ == '__main__':
    theater_search()
